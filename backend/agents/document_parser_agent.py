"""
Feature 008 — Document Parser Agent
Classifies and extracts structured entities from uploaded documents.

VERA_PROFESSIONAL_BOUNDARIES:
    I am your Chief of Staff — not a veterinarian, not your attorney.
    I parse documents for organizational data — team, facilities, schedules.
    I do not interpret clinical records or make diagnostic inferences.

Supported formats: CSV, XLSX/XLS, PDF, PNG/JPG/JPEG/WEBP/GIF/HEIC (OCR)
Pure Python — no external LLM calls.
"""
import csv
import json
import re
import os
import time
from typing import Generator, Optional

VERA_PROFESSIONAL_BOUNDARIES = (
    "I am your Chief of Staff — not a veterinarian, not your attorney. "
    "I parse documents for organizational data — team, facilities, schedules. "
    "I do not interpret clinical records or make diagnostic inferences."
)

# Classification heuristics — column header keywords
_STAFF_KEYWORDS = {"name", "vet", "doctor", "dr", "role", "title", "position",
                   "provider", "physician", "staff", "employee", "dvm", "lvt", "cvt"}
_ROOM_KEYWORDS = {"room", "suite", "exam", "iso", "isolation", "surgical",
                  "treatment", "radiology", "imaging", "ward", "bay", "kennel", "dental"}
_SCHEDULE_KEYWORDS = {"time", "slot", "appointment", "date", "monday", "tuesday",
                      "wednesday", "thursday", "friday", "schedule", "shift"}
_LICENSE_KEYWORDS = {"license", "licence", "dvm", "expires", "expiry", "veterinary",
                     "state", "renewal", "issued", "number"}
_FEE_KEYWORDS = {"price", "fee", "cost", "charge", "rate", "service", "procedure",
                 "amount", "dollars", "$"}

# Role inference keywords
_ROLE_MAP = {
    "dvm": "DVM", "d.v.m": "DVM", "vet": "DVM", "doctor": "DVM",
    "veterinarian": "DVM", "lvt": "LVT", "cvt": "CVT", "tech": "Tech",
    "technician": "Tech", "receptionist": "Receptionist",
    "manager": "Manager", "assistant": "Assistant",
}


def classify_document(path: str, mime_type: str) -> str:
    """
    Classify a document by examining its header row (for spreadsheets)
    or first-page keywords (for PDFs and text).
    Returns: staff_roster | room_list | schedule | license | fee_schedule | unknown
    """
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".csv":
            return _classify_from_headers(_get_csv_headers(path))
        elif ext in (".xlsx", ".xls"):
            return _classify_xlsx(path)
        elif ext == ".pdf":
            return _classify_pdf(path)
        elif ext in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".heic"):
            return "unknown"  # OCR handled at extraction time
    except Exception:
        pass

    return "unknown"


def _get_csv_headers(path: str) -> list:
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
        return [h.lower().strip() for h in headers]
    except Exception:
        return []


def _classify_from_headers(headers: list) -> str:
    header_set = set(h.strip() for h in headers)
    # Flatten multi-word headers into individual tokens
    tokens = set()
    for h in header_set:
        tokens.update(h.split())

    if tokens & _STAFF_KEYWORDS:
        return "staff_roster"
    if tokens & _ROOM_KEYWORDS:
        return "room_list"
    if tokens & _SCHEDULE_KEYWORDS:
        return "schedule"
    if tokens & _LICENSE_KEYWORDS:
        return "license"
    if tokens & _FEE_KEYWORDS:
        return "fee_schedule"
    return "unknown"


def _classify_xlsx(path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        # Check first sheet's first row
        ws = wb.worksheets[0]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c or "").lower().strip() for c in row]
            break
        wb.close()
        return _classify_from_headers(headers)
    except Exception:
        return "unknown"


def _classify_pdf(path: str) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            if pdf.pages:
                text = (pdf.pages[0].extract_text() or "").lower()
                tokens = set(re.findall(r"\w+", text))
                if tokens & _STAFF_KEYWORDS:
                    return "staff_roster"
                if tokens & _LICENSE_KEYWORDS:
                    return "license"
                if tokens & _FEE_KEYWORDS:
                    return "fee_schedule"
    except Exception:
        pass
    return "unknown"


# ------------------------------------------------------------------
# Confidence scoring
# ------------------------------------------------------------------

def assign_confidence(entity_type: str, value: str, context: Optional[dict] = None) -> float:
    """
    Assign confidence score 0.0–1.0 based on entity type and value quality.
    ≥ 0.8 → ✅ high; 0.5–0.79 → ⚠️ medium; < 0.5 → ❓ needs input
    """
    if not value or not str(value).strip():
        return 0.2

    value = str(value).strip()

    if entity_type == "provider":
        # High confidence if matches "Dr. X" or "Name, DVM" pattern
        if re.match(r"Dr\.?\s+[A-Z][a-z]+", value):
            return 0.95
        if re.search(r"\b(DVM|LVT|CVT)\b", value, re.IGNORECASE):
            return 0.85
        if len(value.split()) >= 2:  # First + last name
            return 0.72
        return 0.45  # Single name → needs confirmation

    if entity_type == "room":
        # High if contains a room keyword
        lower = value.lower()
        if any(kw in lower for kw in _ROOM_KEYWORDS):
            return 0.88
        if len(value) == 1 or value.isdigit():
            return 0.3  # "Iso", "1", etc. → low
        return 0.55

    if entity_type == "service":
        return 0.80

    if entity_type == "hour":
        if re.search(r"\d{1,2}(?::\d{2})?(?:\s*[ap]m)?", value, re.IGNORECASE):
            return 0.82
        return 0.50

    if entity_type in ("phone", "address"):
        return 0.78

    return 0.60


# ------------------------------------------------------------------
# Parsers — each yields entity dicts
# ------------------------------------------------------------------

def parse_csv(path: str) -> Generator[dict, None, None]:
    """
    Parse a CSV file. Yields entity dicts.
    Auto-detects column semantics from headers.
    """
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            headers = [h.lower().strip() for h in (reader.fieldnames or [])]
            doc_type = _classify_from_headers(headers)

            for row_num, row in enumerate(reader, start=2):
                entities = _row_to_entities(row, headers, doc_type, row_num, "csv", None)
                for entity in entities:
                    yield entity
    except Exception as e:
        yield {
            "entity_type": "error",
            "source_text": str(e),
            "confidence": 0.0,
            "extracted_fields": {"error": str(e)},
            "source_position": {},
        }


def parse_xlsx(path: str) -> Generator[dict, None, None]:
    """
    Parse an XLSX file, iterating all sheets.
    Yields entity dicts per sheet row.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows_iter = ws.iter_rows(values_only=True)

            # First row = headers
            try:
                header_row = next(rows_iter)
                headers = [str(c or "").lower().strip() for c in header_row]
            except StopIteration:
                continue

            doc_type = _classify_from_headers(headers)

            # Yield a tab_found notification entity
            yield {
                "entity_type": "tab_found",
                "source_text": sheet_name,
                "confidence": 1.0,
                "extracted_fields": {
                    "sheet_name": sheet_name,
                    "doc_type": doc_type,
                    "message": f"Found {sheet_name} tab — reading rows...",
                },
                "source_position": {"sheet": sheet_name, "row": 1},
            }

            for row_num, row_values in enumerate(rows_iter, start=2):
                # Skip empty rows
                if not any(v for v in row_values if v):
                    continue
                row = {h: str(v or "").strip()
                       for h, v in zip(headers, row_values)}
                entities = _row_to_entities(row, headers, doc_type, row_num, "xlsx", sheet_name)
                for entity in entities:
                    yield entity

        wb.close()

    except ImportError:
        yield {
            "entity_type": "error",
            "source_text": "openpyxl not installed",
            "confidence": 0.0,
            "extracted_fields": {"error": "openpyxl required for Excel files"},
            "source_position": {},
        }
    except Exception as e:
        yield {
            "entity_type": "error",
            "source_text": str(e),
            "confidence": 0.0,
            "extracted_fields": {"error": str(e)},
            "source_position": {},
        }


def parse_pdf(path: str) -> Generator[dict, None, None]:
    """
    Parse a PDF using pdfplumber. Yields entity dicts per page.
    Supports staff rosters, license PDFs, and fee schedules.
    """
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                if not text.strip():
                    continue

                # Try table extraction first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        headers = [str(c or "").lower().strip() for c in (table[0] or [])]
                        doc_type = _classify_from_headers(headers)
                        for row_num, row_values in enumerate(table[1:], start=2):
                            row = {h: str(v or "").strip()
                                   for h, v in zip(headers, row_values or [])}
                            entities = _row_to_entities(
                                row, headers, doc_type, row_num, "pdf", f"page_{page_num}"
                            )
                            for entity in entities:
                                yield entity
                else:
                    # Fall back to text extraction
                    for entity in _parse_pdf_text(text, page_num):
                        yield entity

    except ImportError:
        yield {
            "entity_type": "error",
            "source_text": "pdfplumber not installed",
            "confidence": 0.0,
            "extracted_fields": {"error": "pdfplumber required for PDF files"},
            "source_position": {},
        }
    except Exception as e:
        yield {
            "entity_type": "error",
            "source_text": str(e),
            "confidence": 0.0,
            "extracted_fields": {"error": str(e)},
            "source_position": {},
        }


def _parse_pdf_text(text: str, page_num: int) -> Generator[dict, None, None]:
    """Extract entities from raw PDF text (no tables)."""
    from .practice_builder_agent import _extract_providers, _extract_rooms

    providers = _extract_providers(text)
    for i, p in enumerate(providers):
        confidence = assign_confidence("provider", p.get("name", ""))
        yield {
            "entity_type": "provider",
            "source_text": p.get("name", ""),
            "confidence": confidence,
            "extracted_fields": p,
            "source_position": {"page": page_num, "index": i},
        }

    rooms = _extract_rooms(text)
    for i, r in enumerate(rooms):
        confidence = assign_confidence("room", r.get("name", ""))
        yield {
            "entity_type": "room",
            "source_text": r.get("name", ""),
            "confidence": confidence,
            "extracted_fields": r,
            "source_position": {"page": page_num, "index": i},
        }

    # License info
    license_match = re.search(
        r"(?:license|licence)\s*(?:number|#|no\.?)?\s*:?\s*([A-Z0-9-]+)",
        text, re.IGNORECASE
    )
    if license_match:
        yield {
            "entity_type": "license",
            "source_text": license_match.group(0),
            "confidence": 0.80,
            "extracted_fields": {"license_number": license_match.group(1)},
            "source_position": {"page": page_num},
        }


def parse_image(path: str) -> Generator[dict, None, None]:
    """
    OCR an image using pytesseract + Pillow.
    All entities from image OCR have confidence capped at 0.5.
    """
    try:
        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            yield {
                "entity_type": "error",
                "source_text": "OCR libraries not installed",
                "confidence": 0.0,
                "extracted_fields": {"error": "pytesseract and Pillow required for image OCR"},
                "source_position": {},
            }
            return

        img = Image.open(path)
        # Convert HEIC if needed (Pillow-heif required for HEIC)
        text = pytesseract.image_to_string(img)

        if not text.strip():
            yield {
                "entity_type": "error",
                "source_text": "No text detected",
                "confidence": 0.0,
                "extracted_fields": {"error": "OCR found no readable text in this image"},
                "source_position": {},
            }
            return

        # Parse OCR output with capped confidence
        from .practice_builder_agent import _extract_providers, _extract_rooms

        providers = _extract_providers(text)
        for i, p in enumerate(providers):
            yield {
                "entity_type": "provider",
                "source_text": p.get("name", ""),
                "confidence": 0.5,  # OCR confidence cap
                "extracted_fields": p,
                "source_position": {"source": "ocr", "index": i},
            }

        rooms = _extract_rooms(text)
        for i, r in enumerate(rooms):
            yield {
                "entity_type": "room",
                "source_text": r.get("name", ""),
                "confidence": 0.5,
                "extracted_fields": r,
                "source_position": {"source": "ocr", "index": i},
            }

        # Surface raw OCR lines that couldn't be parsed as structured entities
        lines = [l.strip() for l in text.split("\n") if l.strip() and len(l.strip()) > 3]
        if lines and not providers and not rooms:
            yield {
                "entity_type": "raw_text",
                "source_text": text[:500],
                "confidence": 0.3,
                "extracted_fields": {"lines": lines[:20]},
                "source_position": {"source": "ocr"},
            }

    except Exception as e:
        yield {
            "entity_type": "error",
            "source_text": str(e),
            "confidence": 0.0,
            "extracted_fields": {"error": str(e)},
            "source_position": {},
        }


# ------------------------------------------------------------------
# Row → entities mapper
# ------------------------------------------------------------------

def _row_to_entities(row: dict, headers: list, doc_type: str,
                      row_num: int, source_format: str,
                      sheet: Optional[str]) -> list:
    """Convert a spreadsheet row dict to a list of entity dicts."""
    entities = []
    if doc_type in ("staff_roster", "unknown"):
        entity = _row_to_provider(row, headers, row_num, source_format, sheet)
        if entity:
            entities.append(entity)
    if doc_type in ("room_list", "unknown"):
        entity = _row_to_room(row, headers, row_num, source_format, sheet)
        if entity:
            entities.append(entity)
    if doc_type == "fee_schedule":
        entity = _row_to_service(row, headers, row_num, source_format, sheet)
        if entity:
            entities.append(entity)
    return entities


def _row_to_provider(row: dict, headers: list, row_num: int,
                      fmt: str, sheet: Optional[str]) -> Optional[dict]:
    """Attempt to construct a provider entity from a row."""
    # Look for a name column
    name_col = next(
        (h for h in headers if any(kw in h for kw in ("name", "vet", "doctor", "dr", "provider", "employee"))),
        None,
    )
    if not name_col or not row.get(name_col, "").strip():
        return None

    name = row[name_col].strip()
    if not name or name.lower() in ("name", "n/a", "-", ""):
        return None

    # Role
    role_col = next((h for h in headers if any(kw in h for kw in ("role", "title", "position", "type"))), None)
    role_raw = (row.get(role_col, "") or "").strip().lower() if role_col else ""
    role = next((v for k, v in _ROLE_MAP.items() if k in role_raw), "DVM")

    # Working days
    day_col = next((h for h in headers if any(kw in h for kw in ("day", "schedule", "avail", "shift"))), None)
    days_raw = (row.get(day_col, "") or "").strip() if day_col else ""
    days = _parse_days(days_raw)

    # Add Dr. prefix if DVM and not already present
    display_name = name if name.lower().startswith("dr") else f"Dr. {name}" if role == "DVM" else name

    confidence = assign_confidence("provider", display_name)
    return {
        "entity_type": "provider",
        "source_text": f"{name} ({role})" if role else name,
        "confidence": confidence,
        "extracted_fields": {"name": display_name, "role": role, "days": days},
        "source_position": {"sheet": sheet, "row": row_num, "col": name_col},
    }


def _row_to_room(row: dict, headers: list, row_num: int,
                  fmt: str, sheet: Optional[str]) -> Optional[dict]:
    """Attempt to construct a room entity from a row."""
    room_col = next(
        (h for h in headers if any(kw in h for kw in ("room", "suite", "space", "facility", "ward", "bay"))),
        None,
    )
    if not room_col or not row.get(room_col, "").strip():
        return None

    name = row[room_col].strip()
    if not name or name.lower() in ("room", "n/a", "-", ""):
        return None

    confidence = assign_confidence("room", name)
    return {
        "entity_type": "room",
        "source_text": name,
        "confidence": confidence,
        "extracted_fields": {"name": name.title()},
        "source_position": {"sheet": sheet, "row": row_num, "col": room_col},
    }


def _row_to_service(row: dict, headers: list, row_num: int,
                     fmt: str, sheet: Optional[str]) -> Optional[dict]:
    """Attempt to construct a service/fee entity from a row."""
    svc_col = next(
        (h for h in headers if any(kw in h for kw in ("service", "procedure", "description", "item"))),
        None,
    )
    if not svc_col or not row.get(svc_col, "").strip():
        return None

    name = row[svc_col].strip()
    fee_col = next((h for h in headers if any(kw in h for kw in ("fee", "price", "cost", "amount"))), None)
    fee = row.get(fee_col, "") if fee_col else ""

    confidence = assign_confidence("service", name)
    return {
        "entity_type": "service",
        "source_text": name,
        "confidence": confidence,
        "extracted_fields": {"name": name, "fee": fee},
        "source_position": {"sheet": sheet, "row": row_num, "col": svc_col},
    }


def _parse_days(days_str: str) -> list:
    """Parse a days string into a list of abbreviated day names."""
    if not days_str:
        return []

    day_re = re.compile(
        r"\b(mon(?:day)?|tue(?:sday)?|wed(?:nesday)?|thu(?:rsday)?|fri(?:day)?|"
        r"sat(?:urday)?|sun(?:day)?)\b",
        re.IGNORECASE,
    )
    abbrev_map = {
        "mon": "Mon", "monday": "Mon",
        "tue": "Tue", "tuesday": "Tue",
        "wed": "Wed", "wednesday": "Wed",
        "thu": "Thu", "thursday": "Thu",
        "fri": "Fri", "friday": "Fri",
        "sat": "Sat", "saturday": "Sat",
        "sun": "Sun", "sunday": "Sun",
    }
    matches = day_re.findall(days_str)
    return [abbrev_map[m.lower()] for m in matches if m.lower() in abbrev_map]
